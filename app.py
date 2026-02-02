import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime, time
import io
import time # 화면 새로고침용

# ==========================================
# 1. 기본 설정
# ==========================================
FILE_NAME = "2026_2월 급여대장.xlsx"
SHEET_NAME = "근무표(입력)"

st.set_page_config(page_title="OUR 급여관리", layout="wide")
st.title("🍞 OUR 베이커리 급여 입력 (2월)")

# ==========================================
# 2. 엑셀 파일 로드
# ==========================================
if "wb" not in st.session_state:
    try:
        st.session_state.wb = openpyxl.load_workbook(FILE_NAME, data_only=False)
    except:
        st.error(f"폴더에 '{FILE_NAME}' 파일이 없습니다.")
        st.stop()

wb = st.session_state.wb
try:
    sheet = wb[SHEET_NAME]
except:
    st.error(f"엑셀에 '{SHEET_NAME}' 시트가 없습니다.")
    st.stop()

# 화면 표시용 데이터 읽기
@st.cache_data
def load_display_data():
    try:
        return pd.read_excel(FILE_NAME, sheet_name=SHEET_NAME, header=None)
    except:
        return None

df = load_display_data()
if df is None:
    st.error("엑셀 파일을 읽을 수 없습니다.")
    st.stop()

# ==========================================
# 3. 직원 목록 & 날짜 목록 준비
# ==========================================

# 직원 목록 (집계표 Y열)
def get_employees(sheet_obj):
    names = set()
    for r in range(3, 101):
        val = sheet_obj.cell(row=r, column=25).value 
        if val and str(val).strip() != "":
            names.add(str(val))
    return sorted(list(names))

employee_list = get_employees(sheet)

# 날짜 목록
date_row_map = {}
date_options = []

if len(df) > 2:
    for idx, row in df.iloc[2:].iterrows():
        val_date = row[0]
        val_day = row[1]
        
        if pd.notna(val_date):
            d_str = val_date.strftime("%Y-%m-%d") if isinstance(val_date, datetime) else str(val_date).split(" ")[0]
            label = f"{d_str} ({val_day})"
            date_row_map[label] = idx + 1 
            date_options.append(label)

if not date_options:
    st.error("날짜를 찾을 수 없습니다.")
    st.stop()

# ==========================================
# 4. 날짜 선택 및 현재 현황 보여주기 (New!)
# ==========================================
selected_label = st.selectbox("📅 날짜 선택", date_options)
target_row = date_row_map[selected_label]

# --- [여기서부터가 새로 추가된 확인 기능입니다] ---
st.info(f"👇 **{selected_label}** 현재 저장된 근무자 명단입니다. (틀렸다면 아래에서 수정하세요)")

# 현재 엑셀에 저장된 내용을 읽어서 표로 만들기
current_status = []
slot_configs = [
    {"name": "타임 1 (오전)", "col": 3},
    {"name": "타임 2 (미들1)", "col": 7},
    {"name": "타임 3 (미들2)", "col": 11},
    {"name": "타임 4 (오후)", "col": 15},
    {"name": "타임 5 (마감)", "col": 19},
]

for slot in slot_configs:
    base = slot["col"]
    # 엑셀 값 읽기
    c_name = sheet.cell(row=target_row, column=base).value
    c_start = sheet.cell(row=target_row, column=base+1).value
    c_end = sheet.cell(row=target_row, column=base+2).value
    
    if c_name: # 이름이 있는 경우만 표에 추가
        # 시간 예쁘게 표시
        s_str = c_start.strftime("%H:%M") if isinstance(c_start, (datetime, time)) else str(c_start)
        e_str = c_end.strftime("%H:%M") if isinstance(c_end, (datetime, time)) else str(c_end)
        
        current_status.append({
            "근무 타임": slot["name"],
            "이름": c_name,
            "출근": s_str,
            "퇴근": e_str
        })

# 데이터가 있으면 표로 보여주고, 없으면 없다고 표시
if current_status:
    st.dataframe(pd.DataFrame(current_status), use_container_width=True, hide_index=True)
else:
    st.write("🚫 **현재 등록된 근무자가 없습니다.** 아래에서 입력해주세요.")

st.markdown("---")

# ==========================================
# 5. 입력 및 수정 화면
# ==========================================
with st.form("input_form"):
    st.write(f"**✏️ {selected_label} 근무 내용 수정/입력**")
    
    cols = st.columns(2) 
    updates = {}
    new_names_to_add = set()

    for idx, slot in enumerate(slot_configs):
        col_ui = cols[idx % 2]
        base = slot["col"]
        
        # 엑셀 값 읽기 (수정 모드이므로 현재 값을 기본값으로)
        curr_name = sheet.cell(row=target_row, column=base).value
        curr_s = sheet.cell(row=target_row, column=base+1).value
        curr_e = sheet.cell(row=target_row, column=base+2).value
        
        def to_time(v):
            if isinstance(v, datetime): return v.time()
            if isinstance(v, time): return v
            return None

        with col_ui:
            # 값이 있으면 펼쳐서 보여줌
            is_expanded = (curr_name is not None)
            with st.expander(f"{slot['name']}", expanded=is_expanded):
                
                # 이름 선택
                list_opts = ["(직접 입력)"] + employee_list
                def_idx = 0 
                if curr_name and str(curr_name) in employee_list:
                    def_idx = list_opts.index(str(curr_name))
                
                unique_key_sel = f"sel_{base}_{selected_label}"
                unique_key_txt = f"txt_{base}_{selected_label}"
                unique_key_s = f"s_{base}_{selected_label}"
                unique_key_e = f"e_{base}_{selected_label}"

                sel_val = st.selectbox("이름", list_opts, index=def_idx, key=unique_key_sel)
                
                final_n = None
                if sel_val == "(직접 입력)":
                    val_to_show = ""
                    if curr_name and str(curr_name) not in employee_list:
                        val_to_show = str(curr_name)
                    input_name = st.text_input("이름 입력 (비우면 삭제)", value=val_to_show, key=unique_key_txt)
                    
                    if input_name.strip():
                        final_n = input_name.strip()
                        if final_n not in employee_list:
                            new_names_to_add.add(final_n)
                    else:
                        final_n = None
                else:
                    final_n = sel_val

                # 시간 입력
                c1, c2 = st.columns(2)
                new_s = c1.time_input("출근", value=to_time(curr_s), key=unique_key_s)
                new_e = c2.time_input("퇴근", value=to_time(curr_e), key=unique_key_e)
                
                updates[base] = {"n": final_n, "s": new_s, "e": new_e}

    st.markdown("###")
    applied = st.form_submit_button("✅ 저장 및 반영하기", use_container_width=True)

# ==========================================
# 6. 저장 로직 (자동 새로고침 기능 추가)
# ==========================================
if applied:
    # 신규 직원 등록
    if new_names_to_add:
        for new_name in new_names_to_add:
            for r in range(3, 101):
                cell = sheet.cell(row=r, column=25)
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = new_name
                    sheet.cell(row=r, column=24).value = r - 2
                    break
        st.toast(f"🎉 신규 직원이 명단에 등록되었습니다!")

    # 데이터 저장
    for base, data in updates.items():
        if data["n"]:
            sheet.cell(row=target_row, column=base).value = data["n"]
            sheet.cell(row=target_row, column=base+1).value = data["s"]
            sheet.cell(row=target_row, column=base+2).value = data["e"]
        else:
            sheet.cell(row=target_row, column=base).value = None
            sheet.cell(row=target_row, column=base+1).value = None
            sheet.cell(row=target_row, column=base+2).value = None
            
    st.success("저장되었습니다! 화면을 갱신합니다...")
    time.sleep(1) # 1초 대기 후
    st.rerun() # 화면 새로고침! (이제 위쪽 표가 바뀐 내용으로 뜹니다)

# ==========================================
# 7. 다운로드
# ==========================================
st.markdown("---")
output = io.BytesIO()
wb.save(output)
processed_data = output.getvalue()

st.download_button(
    label="📥 엑셀 파일 다운로드 (카톡 전송용)",
    data=processed_data,
    file_name=f"2월급여대장_OUR_{datetime.now().strftime('%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)
